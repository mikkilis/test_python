from openpyxl import workbook,load_workbook

wb = load_workbook('money.xlsx')
ws = wb.active
a1 = ws["A1"]
text = a1.value


wb.save('out_money.xlsx')

print(text)
print("shyam")

print("Done")